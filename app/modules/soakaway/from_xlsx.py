from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook

from app.core.formatters import fmt_number, fmt_date
from app.models.soakaway import empty_soakaway_report


def parse_soakaway_xlsx(path: str | Path):
    path = Path(path)
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    report = empty_soakaway_report()

    report.project["project_name"] = str(ws["B4"].value or "")
    report.project["client"] = str(ws["B5"].value or "")
    report.project["project_engineer"] = str(ws["B6"].value or "")
    report.project["project_id"] = str(ws["B7"].value or "")

    report.test["location_id"] = str(ws["B8"].value or "")
    report.test["test_run"] = str(ws["B9"].value or "1")
    report.test["test_date"] = fmt_date(ws["B10"].value)
    report.test["tested_by"] = str(ws["B11"].value or "")
    report.test["weather"] = str(ws["B15"].value or "")

    report.location["easting"] = fmt_number(ws["B12"].value)
    report.location["northing"] = fmt_number(ws["B13"].value)
    report.location["ground_level"] = fmt_number(ws["B14"].value)

    report.pit["length_m"] = fmt_number(ws["B16"].value)
    report.pit["width_m"] = fmt_number(ws["B17"].value)
    report.pit["depth_m"] = fmt_number(ws["B18"].value)

    report.subfooter["remarks"] = str(ws["B19"].value or "")
    report.subfooter["instrument"] = str(ws["B20"].value or "")
    report.subfooter["methodology"] = str(ws["B21"].value or "BRE 365 Digest rev. 2016")

    report.signoff["originator"] = str(ws["B22"].value or "TK")
    report.signoff["status"] = str(ws["B23"].value or "Prelim")
    report.signoff["checked_approved"] = str(ws["B24"].value or "")
    report.signoff["issue_date"] = fmt_date(ws["B25"].value)

    report.page["fig_no"] = str(ws["B26"].value or "")

    row = 31
    blank = 0
    while row <= 500:
        t = ws[f"A{row}"].value
        d = ws[f"B{row}"].value

        if t is None and d is None:
            blank += 1
            if blank >= 3:
                break
        else:
            blank = 0
            report.readings.append({
                "time_min": fmt_number(t),
                "depth_mbgl": fmt_number(d),
            })

        row += 1

    return report
